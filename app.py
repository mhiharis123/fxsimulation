from flask import Flask, render_template, request, jsonify
import openpyxl
from datetime import datetime
import json

app = Flask(__name__)

class FXProfitCalculator:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.data = self.load_data()
        self.adjustments = {}
    
    def load_data(self):
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb['AM']
        
        daily_data = {}
        current_date = None
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            trade_date = row[0]
            direction = row[9]
            rate = row[10]
            amount = row[11]
            value = row[12]
            
            if direction == 'BUY':
                if trade_date not in daily_data:
                    daily_data[trade_date] = {}
                daily_data[trade_date]['BUY'] = {
                    'rate': rate,
                    'amount': amount,
                    'value': value,
                    'original_rate': rate
                }
            elif direction == 'SELL':
                if trade_date not in daily_data:
                    daily_data[trade_date] = {}
                daily_data[trade_date]['SELL'] = {
                    'rate': rate,
                    'amount': amount,
                    'value': value,
                    'original_rate': rate
                }
            elif direction == 'BANK':
                if trade_date not in daily_data:
                    daily_data[trade_date] = {}
                daily_data[trade_date]['BANK'] = {
                    'rate': rate,
                    'amount': amount,
                    'value': value
                }
        
        return daily_data
    
    def calculate_day_profit(self, date_key, buy_rate_override=None, sell_rate_override=None):
        if date_key not in self.data:
            return None
        
        day = self.data[date_key]
        
        buy_rate = buy_rate_override if buy_rate_override else day['BUY']['original_rate']
        sell_rate = sell_rate_override if sell_rate_override else day['SELL']['original_rate']
        bank_rate = day['BANK']['rate']
        
        buy_amount = day['BUY']['amount']
        sell_amount = day['SELL']['amount']
        bank_amount = day['BANK']['amount']
        
        buy_value = buy_rate * buy_amount
        sell_value = sell_rate * sell_amount
        bank_value = bank_rate * bank_amount
        
        is_sell_trade = sell_value > buy_value
        
        mark_up_raw = (sell_rate - bank_rate) if is_sell_trade else (buy_rate - bank_rate)
        mark_up = abs(mark_up_raw)  # Always show positive value
        
        profit = buy_value - sell_value - bank_value
        
        return {
            'date': date_key.strftime('%Y-%m-%d') if isinstance(date_key, datetime) else date_key,
            'buy_rate': round(buy_rate, 4),
            'sell_rate': round(sell_rate, 4),
            'bank_rate': round(bank_rate, 4),
            'buy_value': round(buy_value, 2),
            'sell_value': round(sell_value, 2),
            'bank_value': round(bank_value, 2),
            'profit': round(profit, 2),
            'trade_type': 'SELL' if is_sell_trade else 'BUY',
            'mark_up': round(mark_up, 4)
        }
    
    def calculate_rates_from_markup(self, date_key, markup_value):
        """Reverse calculate buy/sell rates from desired markup"""
        if date_key not in self.data:
            return None, None
        
        day = self.data[date_key]
        bank_rate = day['BANK']['rate']
        
        # Get original values to determine trade type
        original_buy_value = day['BUY']['value']
        original_sell_value = day['SELL']['value']
        is_sell_trade = original_sell_value > original_buy_value
        
        if is_sell_trade:
            # For SELL: Sell Rate = Bank Rate - Mark Up
            new_sell_rate = bank_rate - markup_value
            return None, new_sell_rate
        else:
            # For BUY: Buy Rate = Bank Rate + Mark Up
            new_buy_rate = bank_rate + markup_value
            return new_buy_rate, None
    
    def calculate_all_profits(self, global_markup=None):
        results = []
        total_profit = 0
        
        for date_key in sorted(self.data.keys()):
            # Format date consistently for lookup
            date_str = date_key.strftime('%Y-%m-%d') if isinstance(date_key, datetime) else date_key
            
            if global_markup is not None:
                # Use global markup override
                new_buy_rate, new_sell_rate = self.calculate_rates_from_markup(date_key, global_markup)
                buy_rate = new_buy_rate if new_buy_rate else self.adjustments.get(date_str, {}).get('buy')
                sell_rate = new_sell_rate if new_sell_rate else self.adjustments.get(date_str, {}).get('sell')
            else:
                # Use individual adjustments
                buy_rate = self.adjustments.get(date_str, {}).get('buy')
                sell_rate = self.adjustments.get(date_str, {}).get('sell')
            
            day_calc = self.calculate_day_profit(date_key, buy_rate, sell_rate)
            if day_calc:
                results.append(day_calc)
                total_profit += day_calc['profit']
        
        return results, total_profit

calculator = FXProfitCalculator('FX October PnL updated.xlsx')
global_markup_override = None

class FXProfitCalculatorPM:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.data = self.load_data()
        self.adjustments = {}
    
    def load_data(self):
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb['PM']
        
        daily_data = {}  # Now stores: {(date, currency, booking_id): {BUY, SELL, BANK, ...}}
        current_currency = None
        current_date = None
        booking_counter = {}  # Track booking numbers per (date, currency)
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            trade_date = row[0]
            currency_col_i = row[8]  # Column I - Primary currency reference
            direction = row[9]
            rate = row[10]
            amount = row[11]
            value = row[12]
            
            # Skip header rows and empty rows (but allow SELL/BANK with None date if direction is valid)
            if not direction:
                continue
            if trade_date is None and direction not in ['BUY', 'SELL', 'BANK']:
                continue
            
            # Skip separator rows
            if isinstance(currency_col_i, str) and '-' in currency_col_i:
                continue
            
            # Detect new booking block when we see a BUY with currency in column I
            if direction == 'BUY' and currency_col_i and isinstance(currency_col_i, str) and currency_col_i not in ['', 'TOTAL PROFIT:']:
                if currency_col_i != current_currency or (trade_date and trade_date != current_date):
                    # New currency or new date = new booking
                    booking_key = (trade_date if trade_date else current_date, currency_col_i)
                    if booking_key not in booking_counter:
                        booking_counter[booking_key] = 0
                    booking_counter[booking_key] += 1
                
                current_currency = currency_col_i
                if trade_date:
                    current_date = trade_date
            
            # Only process BUY, SELL, BANK directions
            if direction not in ['BUY', 'SELL', 'BANK']:
                continue
            
            # Use current date if trade_date is None (for separated currency sections)
            effective_date = trade_date if trade_date else current_date
            
            # Create unique key including booking number
            booking_key = (effective_date, current_currency)
            booking_id = booking_counter.get(booking_key, 1)
            key = (effective_date, current_currency, booking_id)
            
            if key not in daily_data:
                daily_data[key] = {}
            
            if direction == 'BUY':
                daily_data[key]['BUY'] = {
                    'rate': rate,
                    'amount': amount,
                    'value': value,
                    'original_rate': rate
                }
            elif direction == 'SELL':
                daily_data[key]['SELL'] = {
                    'rate': rate,
                    'amount': amount,
                    'value': value,
                    'original_rate': rate
                }
            elif direction == 'BANK':
                daily_data[key]['BANK'] = {
                    'rate': rate,
                    'amount': amount,
                    'value': value
                }
        
        return daily_data
    
    def calculate_rates_from_markup(self, date_key, markup_value):
        """Reverse calculate buy/sell rates from desired markup"""
        if date_key not in self.data:
            return None, None
        
        day = self.data[date_key]
        if 'BANK' not in day:
            return None, None
            
        bank_rate = day['BANK']['rate']
        
        original_buy_value = day.get('BUY', {}).get('value', 0)
        original_sell_value = day.get('SELL', {}).get('value', 0)
        is_sell_trade = original_sell_value > original_buy_value
        
        if is_sell_trade:
            new_sell_rate = bank_rate - markup_value
            return None, new_sell_rate
        else:
            new_buy_rate = bank_rate + markup_value
            return new_buy_rate, None
    
    def calculate_day_profit(self, date_key, buy_rate_override=None, sell_rate_override=None):
        if date_key not in self.data:
            return None
        
        day = self.data[date_key]
        
        if 'BANK' not in day or 'BUY' not in day or 'SELL' not in day:
            return None
        
        buy_rate = buy_rate_override if buy_rate_override else day['BUY']['original_rate']
        sell_rate = sell_rate_override if sell_rate_override else day['SELL']['original_rate']
        bank_rate = day['BANK']['rate']
        
        buy_amount = day['BUY']['amount']
        sell_amount = day['SELL']['amount']
        bank_amount = day['BANK']['amount']
        
        buy_value = buy_rate * buy_amount
        sell_value = sell_rate * sell_amount
        bank_value = bank_rate * bank_amount
        
        is_sell_trade = sell_value > buy_value
        
        mark_up_raw = (sell_rate - bank_rate) if is_sell_trade else (buy_rate - bank_rate)
        mark_up = abs(mark_up_raw)
        
        profit = buy_value - sell_value - bank_value
        
        trade_date = date_key[0]
        currency = date_key[1]
        booking_id = date_key[2] if len(date_key) > 2 else 1
        
        return {
            'date': trade_date.strftime('%Y-%m-%d') if isinstance(trade_date, datetime) else trade_date,
            'currency': currency,
            'booking_id': booking_id,
            'buy_rate': round(buy_rate, 4),
            'sell_rate': round(sell_rate, 4),
            'bank_rate': round(bank_rate, 4),
            'buy_value': round(buy_value, 2),
            'sell_value': round(sell_value, 2),
            'bank_value': round(bank_value, 2),
            'profit': round(profit, 2),
            'trade_type': 'SELL' if is_sell_trade else 'BUY',
            'mark_up': round(mark_up, 4)
        }
    
    def calculate_all_profits(self, global_markup=None):
        results = []
        total_profit = 0
        
        # Sort by date first, then currency, then booking_id
        sorted_keys = sorted([k for k in self.data.keys() if k[0] is not None and k[1] is not None], 
                            key=lambda x: (x[0], x[1], x[2] if len(x) > 2 else 0))
        
        for date_key in sorted_keys:
            date_str = "{}_{}".format(
                date_key[0].strftime('%Y-%m-%d') if isinstance(date_key[0], datetime) else date_key[0],
                date_key[1]
            )
            if len(date_key) > 2:
                date_str = "{}#{}".format(date_str, date_key[2])
            
            if global_markup is not None:
                new_buy_rate, new_sell_rate = self.calculate_rates_from_markup(date_key, global_markup)
                buy_rate = new_buy_rate if new_buy_rate else self.adjustments.get(date_str, {}).get('buy')
                sell_rate = new_sell_rate if new_sell_rate else self.adjustments.get(date_str, {}).get('sell')
            else:
                buy_rate = self.adjustments.get(date_str, {}).get('buy')
                sell_rate = self.adjustments.get(date_str, {}).get('sell')
            
            day_calc = self.calculate_day_profit(date_key, buy_rate, sell_rate)
            if day_calc:
                results.append(day_calc)
                total_profit += day_calc['profit']
        
        return results, total_profit

calculator_pm = FXProfitCalculatorPM('FX October PnL updated.xlsx')
global_markup_override_pm = None

@app.route('/')
def index():
    results, total_profit = calculator.calculate_all_profits()
    return render_template('index.html', results=results, total_profit=total_profit, num_days=len(results))

@app.route('/api/update-rate', methods=['POST'])
def update_rate():
    data = request.json
    date_str = data['date']
    rate_type = data['type']
    rate_value = float(data['rate'])
    
    if date_str not in calculator.adjustments:
        calculator.adjustments[date_str] = {}
    
    if rate_type == 'buy':
        calculator.adjustments[date_str]['buy'] = rate_value
    else:
        calculator.adjustments[date_str]['sell'] = rate_value
    
    results, total_profit = calculator.calculate_all_profits()
    return jsonify({'success': True, 'total_profit': round(total_profit, 2), 'results': results})

@app.route('/api/reset', methods=['POST'])
def reset():
    global global_markup_override
    calculator.adjustments = {}
    global_markup_override = None
    results, total_profit = calculator.calculate_all_profits()
    return jsonify({'success': True, 'total_profit': round(total_profit, 2), 'results': results})

@app.route('/api/set-global-markup', methods=['POST'])
def set_global_markup():
    global global_markup_override
    data = request.json
    markup_value = data.get('markup')
    
    if markup_value is not None:
        global_markup_override = float(markup_value)
        results, total_profit = calculator.calculate_all_profits(global_markup=global_markup_override)
    else:
        global_markup_override = None
        results, total_profit = calculator.calculate_all_profits()
    
    return jsonify({'success': True, 'total_profit': round(total_profit, 2), 'results': results, 'global_markup': global_markup_override})

@app.route('/api/get-data')
def get_data():
    results, total_profit = calculator.calculate_all_profits()
    return jsonify({'results': results, 'total_profit': round(total_profit, 2)})

@app.route('/pm')
def pm_sheet():
    results, total_profit = calculator_pm.calculate_all_profits()
    currencies = list(set([r['currency'] for r in results]))
    return render_template('pm.html', results=results, total_profit=total_profit, num_rows=len(results), currencies=sorted(currencies))

@app.route('/api/pm/update-rate', methods=['POST'])
def pm_update_rate():
    global global_markup_override_pm
    data = request.json
    date_str = data['date']
    currency = data['currency']
    booking_id = data.get('booking_id', 1)
    rate_type = data['type']
    rate_value = float(data['rate'])
    
    key = "{}_{}#{}".format(date_str, currency, booking_id)
    
    if key not in calculator_pm.adjustments:
        calculator_pm.adjustments[key] = {}
    
    if rate_type == 'buy':
        calculator_pm.adjustments[key]['buy'] = rate_value
    else:
        calculator_pm.adjustments[key]['sell'] = rate_value
    
    global_markup_override_pm = None
    results, total_profit = calculator_pm.calculate_all_profits()
    return jsonify({'success': True, 'total_profit': round(total_profit, 2), 'results': results})

@app.route('/api/pm/reset', methods=['POST'])
def pm_reset():
    global global_markup_override_pm
    calculator_pm.adjustments = {}
    global_markup_override_pm = None
    results, total_profit = calculator_pm.calculate_all_profits()
    return jsonify({'success': True, 'total_profit': round(total_profit, 2), 'results': results})

@app.route('/api/pm/set-global-markup', methods=['POST'])
def pm_set_global_markup():
    global global_markup_override_pm
    data = request.json
    markup_value = data.get('markup')
    
    if markup_value is not None:
        global_markup_override_pm = float(markup_value)
        results, total_profit = calculator_pm.calculate_all_profits(global_markup=global_markup_override_pm)
    else:
        global_markup_override_pm = None
        results, total_profit = calculator_pm.calculate_all_profits()
    
    return jsonify({'success': True, 'total_profit': round(total_profit, 2), 'results': results, 'global_markup': global_markup_override_pm})

if __name__ == '__main__':
    import os
    debug_mode = os.getenv('FLASK_DEBUG', '0') == '1'
    port = int(os.getenv('PORT', 5000))
    host = os.getenv('HOST', '0.0.0.0')
    app.run(debug=debug_mode, port=port, host=host)
