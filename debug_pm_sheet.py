import openpyxl
from datetime import datetime

file_path = 'FX October PnL updated.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['PM']

print("Analyzing PM Sheet structure:")
print("=" * 80)

# Examine rows
print("\nFirst 15 rows with data details:")
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=15, values_only=True), 2):
    if any(cell for cell in row if cell is not None):
        print(f"\nRow {i}: {row}")
        trade_date = row[0]
        currency_pair = row[8]
        direction = row[9]
        print(f"  -> Date: {trade_date}, Currency: {currency_pair}, Direction: {direction}")

print("\n" + "=" * 80)
print("\nChecking for BUY/SELL/BANK directions:")

buy_count = 0
sell_count = 0
bank_count = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    direction = row[9]
    if direction == 'BUY':
        buy_count += 1
    elif direction == 'SELL':
        sell_count += 1
    elif direction == 'BANK':
        bank_count += 1

print(f"BUY entries: {buy_count}")
print(f"SELL entries: {sell_count}")
print(f"BANK entries: {bank_count}")
