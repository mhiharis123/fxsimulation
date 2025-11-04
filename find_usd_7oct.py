import openpyxl
from datetime import datetime

file_path = 'FX October PnL updated.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['PM']

print("Searching for USD/MYR entries on all dates")
print("=" * 100)

target_date = datetime(2025, 10, 7)

# Find all rows mentioning USD/MYR or "USD"
usd_rows = []

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
    # Check all columns for USD
    for j, cell in enumerate(row):
        if cell and isinstance(cell, str) and 'USD' in cell:
            trade_date = row[0]
            currency_col8 = row[8]
            direction = row[9]
            rate = row[10]
            
            usd_rows.append({
                'row': i,
                'date': trade_date,
                'col': j,
                'value': cell,
                'col8': currency_col8,
                'direction': direction,
                'rate': rate
            })

print(f"Found {len(usd_rows)} rows with USD mention:")
for entry in usd_rows[:30]:  # Show first 30
    print(f"  Row {entry['row']}: {entry['date']} - Col{entry['col']}: {entry['value']} | Col8: {entry['col8']} | Direction: {entry['direction']}")

# Group by date
print("\n" + "=" * 100)
print("USD/MYR entries grouped by date:")
dates_with_usd = {}
for entry in usd_rows:
    if isinstance(entry['date'], datetime):
        date_key = entry['date'].date()
        if date_key not in dates_with_usd:
            dates_with_usd[date_key] = []
        dates_with_usd[date_key].append(entry)

for date_key in sorted(dates_with_usd.keys()):
    print(f"\n{date_key}:")
    for entry in dates_with_usd[date_key]:
        print(f"  Row {entry['row']}: {entry['value']} (Direction: {entry['direction']}, Rate: {entry['rate']})")
