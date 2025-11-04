import openpyxl
import json

file_path = 'FX October PnL updated.xlsx'
wb = openpyxl.load_workbook(file_path)
print("Sheet names:", wb.sheetnames)

# Check AM sheet
if 'AM' in wb.sheetnames:
    ws = wb['AM']
    print(f"\nAM Sheet - Rows: {ws.max_row}, Columns: {ws.max_column}")
    print("\nFirst 5 rows:")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        print(f"Row {i}: {row}")
