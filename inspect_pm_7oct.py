import openpyxl
from datetime import datetime

file_path = 'FX October PnL updated.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['PM']

print("Inspecting PM Sheet around 2025-10-07")
print("=" * 100)

# Find rows with 2025-10-07
target_date = datetime(2025, 10, 7)

for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    trade_date = row[0]
    
    # Check if it's the target date
    if isinstance(trade_date, datetime) and trade_date.date() == target_date.date():
        currency_col8 = row[8]
        direction = row[9]
        rate = row[10]
        amount = row[11]
        value = row[12]
        
        print(f"Row {i}: Date={trade_date.date()}, Col8={currency_col8}, Direction={direction}, Rate={rate}, Amount={amount}, Value={value}")

print("\n" + "=" * 100)
print("Looking for all unique currencies in PM sheet:")

currencies_found = {}
current_currency = None

for row in ws.iter_rows(min_row=2, values_only=True):
    trade_date = row[0]
    currency_pair = row[8]
    direction = row[9]
    
    if not direction or trade_date is None:
        continue
    
    # Update current currency if present
    if currency_pair and currency_pair not in [None, '------', '']:
        current_currency = currency_pair
    
    if direction in ['BUY', 'SELL', 'BANK']:
        if isinstance(trade_date, datetime):
            date_key = trade_date.date()
            if date_key not in currencies_found:
                currencies_found[date_key] = set()
            if current_currency:
                currencies_found[date_key].add(current_currency)

# Show USD/MYR entries
print("\nDates with USD/MYR currency:")
for date_key in sorted(currencies_found.keys()):
    if 'USD/MYR' in currencies_found[date_key]:
        print(f"  {date_key}: {sorted(currencies_found[date_key])}")
