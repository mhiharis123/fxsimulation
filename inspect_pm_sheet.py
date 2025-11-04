import openpyxl
import json

file_path = 'FX October PnL updated.xlsx'
wb = openpyxl.load_workbook(file_path)

if 'PM' in wb.sheetnames:
    ws = wb['PM']
    print(f"PM Sheet - Rows: {ws.max_row}, Columns: {ws.max_column}")
    print("\nFirst 10 rows:")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        print(f"Row {i}: {row}")
    
    print("\n\nChecking unique currencies and dates:")
    currencies = set()
    dates = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Trade Date
            dates.add(row[0])
        if row[2]:  # CCY Pair
            currencies.add(row[2])
    
    print(f"Unique Dates: {sorted(dates)}")
    print(f"Unique Currencies: {sorted(currencies)}")
