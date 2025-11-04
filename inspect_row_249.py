import openpyxl

file_path = 'FX October PnL updated.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['PM']

print("Inspecting Row 249 in detail")
print("=" * 100)

row = ws[249]
print(f"\nRow 249 contents:")
for col_idx, cell in enumerate(row, 1):
    print(f"  Col {col_idx}: {cell.value}")

print("\n" + "=" * 100)
print("Looking at rows 248-252:")
for row_num in range(248, 253):
    row = ws[row_num]
    print(f"\nRow {row_num}:")
    print(f"  Col 1 (A): {row[0].value}")
    print(f"  Col 3 (C): {row[2].value}")
    print(f"  Col 9 (I): {row[8].value}")
    print(f"  Col 10 (J): {row[9].value}")
    print(f"  Col 11 (K): {row[10].value}")
